import pandas as pd
import re
from openpyxl import load_workbook

def filter_columns_by_prefix(dataframe, prefixes):
    return [col for col in dataframe.columns if any(col.startswith(prefix) for prefix in prefixes)]

def scrupdate(url, sheetName, excel_file, stathead_file_path, home_team, away_team):
    # 1. Scrape TeamRankings Data
    tables = pd.read_html(url)
    df = tables[0]

    if sheetName == "SOS" and set(["Team", "Rating"]).issubset(df.columns):
        # Filter and clean SOS data
        df_filtered = df[["Team", "Rating"]].copy()
        df_filtered.rename(columns={"Rating": "SOS"}, inplace=True)
        df_filtered["Team"] = df_filtered["Team"].str.replace(r"\s\(\d+-\d+\)", "", regex=True)
        sos_dict = dict(zip(df_filtered["Team"].str.lower(), df_filtered["SOS"]))

    elif sheetName == "spreadmovhome" and set(["H/A/N", "LAL Line", "Result"]).issubset(df.columns):
        # Extract spread movement for home team
        df_filtered = df[["H/A/N", "LAL Line", "Result"]].copy()
        df_filtered["Result"] = df_filtered["Result"].apply(
            lambda x: int(re.sub(r'W by (\d+)', r'\1', x))
            if 'W by' in x
            else -int(re.sub(r'L by (\d+)', r'\1', x))
        )
        df_filtered.rename(columns={"Result": "home_spread_movement"}, inplace=True)

    elif sheetName == "spreadmovaway" and set(["H/A/N", "OKC Line", "Result"]).issubset(df.columns):
        # Extract spread movement for away team
        df_filtered = df[["H/A/N", "OKC Line", "Result"]].copy()
        df_filtered["Result"] = df_filtered["Result"].apply(
            lambda x: int(re.sub(r'W by (\d+)', r'\1', x))
            if 'W by' in x
            else -int(re.sub(r'L by (\d+)', r'\1', x))
        )
        df_filtered.rename(columns={"Result": "away_spread_movement"}, inplace=True)

    else:
        df_filtered = df.copy()

    # 2. Clean Stathead Data
    stathead_data = pd.read_excel(stathead_file_path, skiprows=1)

    # Strip any leading or trailing whitespace from column names
    stathead_data.columns = stathead_data.columns.str.strip()

    # Define prefixes for filtering columns
    team_prefixes = ['2P', '3P', 'FT', 'ORB', 'DRB', 'TOV', 'TRB']
    opponent_prefixes = ['2P.1', '3P.1', 'FT.1', 'ORB.1', 'DRB.1', 'TOV.1', 'STL']

    # Filter columns using prefixes
    team_stats_cols = filter_columns_by_prefix(stathead_data, team_prefixes)
    opponent_stats_cols = filter_columns_by_prefix(stathead_data, opponent_prefixes)

    # Ensure columns exist before extracting
    try:
        filtered_data = stathead_data[team_stats_cols + opponent_stats_cols].copy()
    except KeyError as e:
        print(f"KeyError: {e}. Check column names!")
        filtered_data = pd.DataFrame()  # Empty DataFrame as a fallback

    # Define rename mapping and apply
    rename_map = {
        "2P%": "2ptpercent", "3P%": "3ptpercent", "FT%": "ftpercent",
        "ORB%": "orbpercent", "DRB%": "drbpercent", "TOV%": "tov",
        "2P%.1": "op2ptpercent", "3P%.1": "op3ptpercent", "FT%.1": "opftpercent",
        "ORB%.1": "oporpercent", "DRB%.1": "opdrpercent", "TOV%.1": "optov", "STL%": "opstealpercent",
        "TRB%": "totalreboundpercent", "TRB": "totalrebound"
    }
    filtered_data.rename(columns=rename_map, inplace=True)

    # Remove columns not listed in rename_map
    filtered_data = filtered_data.loc[:, filtered_data.columns.isin(rename_map.values())]

    # Drop duplicate columns, keeping the first instance
    filtered_data = filtered_data.loc[:, ~filtered_data.columns.duplicated()]

    # 3. Append SOS Columns
    if sheetName == "SOS":
        filtered_data["home_team_sos"] = sos_dict.get(home_team.lower(), None)
        filtered_data["away_team_sos"] = sos_dict.get(away_team.lower(), None)

    # Combine Cleaned Data
    combined_data = pd.concat([filtered_data, df_filtered], axis=1)

    # Export to Excel
    with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a") as writer:
        combined_data.to_excel(writer, sheet_name=sheetName, index=False)

    print(f"Data updated in {sheetName} sheet.")

def append_columns_to_new_sheet(excel_file, base_sheet_name, additional_sheet_names, new_sheet_name):
    """
    Appends the last three columns from additional sheets to the base sheet
    and saves the result to a new sheet in the same Excel file.
    
    Parameters:
    - excel_file: Path to the Excel file.
    - base_sheet_name: Name of the base sheet.
    - additional_sheet_names: List of sheet names to extract last three columns from.
    - new_sheet_name: Name of the new sheet to save the updated data.
    """
    # Load base dataframe
    base_df = pd.read_excel(excel_file, sheet_name=base_sheet_name)

    # Load workbook
    wb = load_workbook(excel_file)

    # Append data from additional sheets
    for sheet_name in additional_sheet_names:
        temp_df = pd.read_excel(excel_file, sheet_name=sheet_name)
        print(f"Appending columns from {sheet_name}:")
        base_df = pd.concat([base_df, temp_df.iloc[:, -3:]], axis=1)

    # Save the updated dataframe to a new sheet
    with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        base_df.to_excel(writer, sheet_name=new_sheet_name, index=False)

    print(f"Data saved to a new sheet: {new_sheet_name}.")

    wb = load_workbook(excel_file)

    sheets_to_keep = {"hometeam_data", "awayteam_data", new_sheet_name}
    for sheet in wb.sheetnames:
        if sheet not in sheets_to_keep:
            del wb[sheet]
            print(f"Deleted sheet: {sheet}")
    
    wb.save(excel_file)

# Define the Excel file path
excel_file = './Masterfile.xlsx'
stathead_file_paths = [
    r'C:\Users\Massimo Camuso\Desktop\NBA-NCAAB Model\teamdata\lakersdata.xlsx',  # Home team path
    r"C:\Users\Massimo Camuso\Desktop\NBA-NCAAB Model\teamdata\okcdata.xlsx"]

# Define the home and away teams
teams = ["LA Lakers", "Okla City"]

# Define the stats information with URLs and sheet names
stats_info = [
    {"url": "https://www.teamrankings.com/nba/ranking/schedule-strength-by-other?date=2024-12-28", "sheetName": "SOS"},
    {"url": "https://www.teamrankings.com/nba/team/los-angeles-lakers/ats-results", "sheetName": "spreadmovhome"},
    {"url": "https://www.teamrankings.com/nba/team/oklahoma-city-thunder/ats-results", "sheetName": "spreadmovaway"}
]

new_sheet_names = ["hometeam_data", "awayteam_data"]

# Iterate through the stathead file paths and corresponding teams
for i, path in enumerate(stathead_file_paths):
    # Get the corresponding team and new sheet name
    team = teams[i]
    new_sheet_name = new_sheet_names[i]

    # Process the stathead data for the current team
    for stat_info in stats_info:
        scrupdate(stat_info['url'], stat_info['sheetName'], excel_file, path, teams[0], teams[1])

    # Append columns to the new sheet for the current team
    append_columns_to_new_sheet(
        excel_file=excel_file,
        base_sheet_name="SOS",
        additional_sheet_names=["spreadmovhome", "spreadmovaway"],
        new_sheet_name=new_sheet_name
    )

